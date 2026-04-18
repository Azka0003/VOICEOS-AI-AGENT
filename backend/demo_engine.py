"""
demo_engine.py — DebtPilot Demo Engine

PURPOSE:
  Makes the system feel live even with static data.
  On every backend startup, injects 1–2 new invoice entries with
  fresh due dates and varied next_actions so the agent pipeline
  always has something new to process and demonstrate.

HOW IT WORKS:
  1. Reads existing invoices from Excel to avoid duplicates
  2. Picks a random scenario from DEMO_SCENARIOS
  3. Injects it into BOTH mock_invoices.json AND invoices.xlsx
  4. The next batch run will pick it up and the agent will act on it

DEMO SCENARIOS are designed to cover every agent path:
  - email (friendly, urgent, final notice)
  - call trigger
  - HITL trigger (missing contact)
  - escalation

WHY THIS APPROACH:
  The due_date is set dynamically relative to today, so days_overdue
  is always real. Each startup adds a new client cycle, making
  the lineage log grow and the dashboard feel active.
"""

import os
import json
import random
import openpyxl
from datetime import date, datetime, timedelta, timezone

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
EXCEL_PATH = os.path.join(DATA_DIR, "invoices.xlsx")
JSON_PATH = os.path.join(DATA_DIR, "mock_invoices.json")
INJECTED_LOG = os.path.join(DATA_DIR, "demo_injected.json")

# ── Demo scenario templates ───────────────────────────────────────────────────
# Each scenario maps to a different agent action path.
# due_offset = how many days AGO the invoice was due (negative = overdue)

DEMO_SCENARIOS = [
    {
        "label": "email_friendly",
        "client": "Sunrise Retail",
        "contact_name": "Ananya Sharma",
        "contact_email": "farazstudy112@gmail.com",   # real inbox for demo
        "contact_phone": None,
        "amount_range": (8000, 25000),
        "due_offset_days": -18,          # 18 days overdue → friendly reminder
        "risk_score": 28,
        "risk_label": "Low",
        "dispute_flag": False,
        "next_action": "send_friendly_reminder",
        "status": "overdue",
        "payment_history": ["Paid INV-2025-03 on time", "First overdue invoice"],
    },
    {
        "label": "email_urgent",
        "client": "Patel Industries",
        "contact_name": "Kiran Patel",
        "contact_email": "farazstudy112@gmail.com",
        "contact_phone": None,
        "amount_range": (35000, 80000),
        "due_offset_days": -38,          # 38 days overdue → urgent followup
        "risk_score": 55,
        "risk_label": "Medium",
        "dispute_flag": False,
        "next_action": "send_urgent_followup",
        "status": "overdue",
        "payment_history": ["Late on INV-2025-01", "No response to first reminder"],
    },
    {
        "label": "email_final",
        "client": "Gupta Wholesale",
        "contact_name": "Deepak Gupta",
        "contact_email": "farazstudy112@gmail.com",
        "contact_phone": None,
        "amount_range": (90000, 200000),
        "due_offset_days": -65,          # 65 days overdue → final notice
        "risk_score": 72,
        "risk_label": "High",
        "dispute_flag": False,
        "next_action": "send_final_notice",
        "status": "overdue",
        "payment_history": ["Missed 2 reminders", "Partial payment only on prior invoice"],
    },
    {
        "label": "call_trigger",
        "client": "Verma Exports",
        "contact_name": "Suresh Verma",
        "contact_email": "farazstudy112@gmail.com",
        "contact_phone": "+919634143593",   # real number for demo
        "amount_range": (120000, 300000),
        "due_offset_days": -72,          # 72 days overdue → call
        "risk_score": 82,
        "risk_label": "High",
        "dispute_flag": False,
        "next_action": "schedule_call",
        "status": "overdue",
        "payment_history": ["Silent for 60+ days", "Prior call went unanswered", "High-value account"],
    },
    {
        "label": "hitl_missing_contact",
        "client": "Khan & Brothers",
        "contact_name": None,            # deliberately missing → HITL
        "contact_email": None,
        "contact_phone": None,
        "amount_range": (50000, 150000),
        "due_offset_days": -45,
        "risk_score": 61,
        "risk_label": "Medium",
        "dispute_flag": False,
        "next_action": "resolve_contact_details",
        "status": "overdue",
        "payment_history": ["Contact details need verification", "Invoice sent to registered address"],
    },
    {
        "label": "email_dispute",
        "client": "Reddy Logistics",
        "contact_name": "Priya Reddy",
        "contact_email": "farazstudy112@gmail.com",
        "contact_phone": None,
        "amount_range": (40000, 100000),
        "due_offset_days": -55,
        "risk_score": 88,
        "risk_label": "High",
        "dispute_flag": True,
        "next_action": "send_urgent_followup",
        "status": "overdue",
        "payment_history": ["Dispute raised on delivery quality", "Pending resolution since 30 days"],
    },
]

# ── Already-seen client tracker ───────────────────────────────────────────────

def _get_injected_clients() -> set:
    """Returns set of clients already injected in prior runs."""
    if not os.path.exists(INJECTED_LOG):
        return set()
    try:
        with open(INJECTED_LOG) as f:
            return set(json.load(f))
    except Exception:
        return set()


def _mark_injected(client: str):
    """Records that this client was injected so we don't duplicate."""
    existing = _get_injected_clients()
    existing.add(client)
    with open(INJECTED_LOG, "w") as f:
        json.dump(list(existing), f)


def _get_existing_clients_from_excel() -> set:
    """Reads current Excel to find all client names already there."""
    if not os.path.exists(EXCEL_PATH):
        return set()
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        clients = set()
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=2).value
            if val:
                clients.add(str(val).strip())
        return clients
    except Exception:
        return set()


def _next_invoice_id(existing_json: list) -> str:
    """Generates next INV### id that doesn't clash."""
    existing_ids = {item.get("id", "") for item in existing_json}
    for n in range(100, 999):
        candidate = f"INV{n:03d}"
        if candidate not in existing_ids:
            return candidate
    return f"INV{random.randint(1000, 9999)}"


# ── Core injection logic ──────────────────────────────────────────────────────

def inject_demo_entries(count: int = 2) -> list[dict]:
    """
    Picks `count` unused scenarios and injects them into both data stores.
    Returns list of injected scenario dicts (for logging).
    Returns empty list if all scenarios already used.
    """
    already_injected = _get_injected_clients()
    existing_excel_clients = _get_existing_clients_from_excel()

    # Find unused scenarios
    available = [
        s for s in DEMO_SCENARIOS
        if s["client"] not in already_injected
        and s["client"] not in existing_excel_clients
    ]

    if not available:
        print("[DEMO ENGINE] All demo scenarios already injected. Resetting tracker.")
        # Reset so we can run the demo again
        if os.path.exists(INJECTED_LOG):
            os.remove(INJECTED_LOG)
        available = [
            s for s in DEMO_SCENARIOS
            if s["client"] not in existing_excel_clients
        ]

    if not available:
        print("[DEMO ENGINE] Excel already has all demo clients. Skipping injection.")
        return []

    # Pick scenarios
    chosen = random.sample(available, min(count, len(available)))
    injected = []

    # Load current JSON
    try:
        with open(JSON_PATH) as f:
            json_data = json.load(f)
    except Exception:
        json_data = []

    # Load Excel
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        excel_max_row = ws.max_row
    except Exception as e:
        print(f"[DEMO ENGINE] Could not open Excel: {e}")
        return []

    today = date.today()

    for scenario in chosen:
        inv_id = _next_invoice_id(json_data)
        amount = random.randint(*scenario["amount_range"])
        due_date = today + timedelta(days=scenario["due_offset_days"])
        days_overdue = abs(scenario["due_offset_days"])

        # ── Inject into JSON ──────────────────────────────────────────────────
        json_entry = {
            "id": inv_id,
            "client": scenario["client"],
            "amount": amount,
            "due_date": due_date.isoformat(),
            "days_overdue": days_overdue,
            "status": scenario["status"],
            "contact_name": scenario["contact_name"] or "",
            "contact_email": scenario["contact_email"] or "",
            "contact_phone": scenario["contact_phone"] or "",
            "payment_history": scenario["payment_history"],
            "dispute_flag": scenario["dispute_flag"],
            "risk_score": scenario["risk_score"],
            "risk_label": scenario["risk_label"],
            "next_action": scenario["next_action"],
        }
        json_data.append(json_entry)

        # ── Inject into Excel ─────────────────────────────────────────────────
        excel_max_row += 1
        now_str = datetime.now(timezone.utc).isoformat()

        ws.cell(row=excel_max_row, column=1).value  = inv_id                        # Invoice ID
        ws.cell(row=excel_max_row, column=2).value  = scenario["client"]            # Client
        ws.cell(row=excel_max_row, column=3).value  = amount                        # Amount
        ws.cell(row=excel_max_row, column=4).value  = due_date.isoformat()          # Due Date
        ws.cell(row=excel_max_row, column=5).value  = days_overdue                  # Days Overdue
        ws.cell(row=excel_max_row, column=6).value  = scenario["status"]            # Status
        ws.cell(row=excel_max_row, column=7).value  = scenario["contact_name"] or "" # Contact Name
        ws.cell(row=excel_max_row, column=8).value  = scenario["contact_email"] or "" # Contact Email
        ws.cell(row=excel_max_row, column=9).value  = scenario["risk_score"]        # Risk Score
        ws.cell(row=excel_max_row, column=10).value = scenario["risk_label"]        # Risk Label
        ws.cell(row=excel_max_row, column=11).value = "yes" if scenario["dispute_flag"] else "no"  # Dispute
        ws.cell(row=excel_max_row, column=12).value = "Never"                       # Last Contact Date
        ws.cell(row=excel_max_row, column=13).value = "none"                        # Last Contact Type
        ws.cell(row=excel_max_row, column=14).value = 0                             # Contact Count
        ws.cell(row=excel_max_row, column=15).value = scenario["next_action"]       # Next Action ← KEY
        ws.cell(row=excel_max_row, column=16).value = "demo_engine"                 # Last Updated By
        ws.cell(row=excel_max_row, column=17).value = now_str                       # Last Updated At

        _mark_injected(scenario["client"])
        injected.append({"id": inv_id, "client": scenario["client"], "label": scenario["label"], "next_action": scenario["next_action"]})
        print(f"[DEMO ENGINE] Injected: {inv_id} | {scenario['client']} | {scenario['next_action']} | ₹{amount:,}")

    # Save both files
    try:
        with open(JSON_PATH, "w") as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        wb.save(EXCEL_PATH)
        print(f"[DEMO ENGINE] Saved {len(injected)} entries to Excel + JSON.")
    except Exception as e:
        print(f"[DEMO ENGINE] Save failed: {e}")
        return []

    return injected


# ── Entry point for direct test ───────────────────────────────────────────────
if __name__ == "__main__":
    result = inject_demo_entries(count=2)
    print(f"\nInjected {len(result)} demo entries:")
    for r in result:
        print(f"  {r}")
