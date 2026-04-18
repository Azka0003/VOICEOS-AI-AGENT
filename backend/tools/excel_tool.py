import os
import json
import fcntl
import openpyxl
from datetime import date, datetime, timezone
from contextlib import contextmanager
from openpyxl.styles import PatternFill, Font

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIGURATION & CONSTANTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data")
EXCEL_PATH = os.path.join(DATA_DIR, "invoices.xlsx")
JSON_PATH = os.path.join(DATA_DIR, "mock_invoices.json")
LOG_PATH = os.path.join(DATA_DIR, "lineage_log.json")

# Formatting
FILL_HEADER = PatternFill("solid", fgColor="2F4F8F")
FONT_HEADER = Font(bold=True, color="FFFFFF")

HEADERS =[
    "Invoice ID", "Client", "Amount (₹)", "Due Date", "Days Overdue",
    "Status", "Contact Name", "Contact Email", "Risk Score", "Risk Label",
    "Dispute", "Last Contact Date", "Last Contact Type", "Contact Count",
    "Next Action", "Last Updated By", "Last Updated At"
]

PROTECTED_ACTIONS =["escalate_to_legal", "disputed_under_review"]

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# LOCKING MECHANISM & HELPER FUNCTIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@contextmanager
def locked_excel(filepath: str, mode: str = "r"):
    """
    Opens the Excel file with an OS-level file lock to ensure concurrency safety.
    mode: "r" for read-only, "w" for read-write.
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    lock_path = filepath + ".lock"
    
    with open(lock_path, "w") as lock_file:
        fcntl.flock(lock_file, fcntl.LOCK_EX)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            yield wb
            if mode == "w":
                wb.save(filepath)
        finally:
            fcntl.flock(lock_file, fcntl.LOCK_UN)

def _log_lineage(entry: dict):
    """Safely append to the lineage log."""
    os.makedirs(os.path.dirname(LOG_PATH), exist_ok=True)
    try:
        with open(LOG_PATH, "r") as f:
            logs = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        logs = []
    
    entry["timestamp"] = datetime.now(timezone.utc).isoformat()
    logs.append(entry)
    
    with open(LOG_PATH, "w") as f:
        json.dump(logs, f, indent=2)

def recalculate_days_overdue(due_date_raw) -> int:
    """Calculates days overdue accurately on every read."""
    if not due_date_raw:
        return 0
        
    try:
        # Handle datetime objects natively returned by openpyxl
        if hasattr(due_date_raw, "date"):
            due = due_date_raw.date()
        # Handle string parsing for DD-MMM-YYYY
        elif isinstance(due_date_raw, str):
            due = datetime.strptime(due_date_raw.strip(), "%d-%b-%Y").date()
        else:
            return 0
            
        today = date.today()
        delta = (today - due).days
        return max(0, delta)
    except Exception:
        return 0

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL TOOL CLASS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

class ExcelTool:
    """
    The Bookkeeper and Source of Truth for DebtPilot.
    """

    def __init__(self, filepath=EXCEL_PATH):
        self.filepath = filepath
        self._init_sheet()

    def _init_sheet(self):
        """Creates and formats the file if it doesn't exist."""
        if not os.path.exists(self.filepath):
            os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices"
            ws.append(HEADERS)
            for col_num in range(1, 18):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
            wb.save(self.filepath)

    def get_all_invoices(self) -> list[dict]:
        """
        Returns all rows (excluding 'paid') with dynamically recalculated days_overdue.
        Automatically fixes stale overdue values in the Excel file.
        """
        invoices =[]
        updates_made = False

        with locked_excel(self.filepath, mode="w") as wb:
            ws = wb.active
            
            for row in range(2, ws.max_row + 1):
                status = str(ws.cell(row=row, column=6).value or "").strip().lower()
                if status == "paid":
                    continue

                inv_id = str(ws.cell(row=row, column=1).value or "").strip()
                if not inv_id:
                    continue

                due_date_raw = ws.cell(row=row, column=4).value
                old_overdue = int(ws.cell(row=row, column=5).value or 0)
                
                # Recalculate days overdue
                real_overdue = recalculate_days_overdue(due_date_raw)
                
                if real_overdue != old_overdue:
                    ws.cell(row=row, column=5).value = real_overdue
                    updates_made = True

                invoices.append({
                    "id": inv_id,
                    "client": str(ws.cell(row=row, column=2).value or "").strip(),
                    "amount": int(ws.cell(row=row, column=3).value or 0),
                    "due_date": str(due_date_raw)[:10] if due_date_raw else "",
                    "days_overdue": real_overdue,
                    "status": status,
                    "contact_name": str(ws.cell(row=row, column=7).value or "").strip(),
                    "contact_email": str(ws.cell(row=row, column=8).value or "").strip(),
                    "risk_score": int(ws.cell(row=row, column=9).value or 0),
                    "risk_label": str(ws.cell(row=row, column=10).value or "").strip(),
                    "dispute_flag": str(ws.cell(row=row, column=11).value or "").strip().lower() == "yes",
                    "last_contact_date": str(ws.cell(row=row, column=12).value or "Never"),
                    "last_contact_type": str(ws.cell(row=row, column=13).value or "none"),
                    "contact_count": int(ws.cell(row=row, column=14).value or 0),
                    "next_action": str(ws.cell(row=row, column=15).value or "").strip(),
                    "last_updated_by": str(ws.cell(row=row, column=16).value or ""),
                    "last_updated_at": str(ws.cell(row=row, column=17).value or "")
                })

            # Revert to read-only mode if no updates occurred to prevent unnecessary saves
            if not updates_made:
                wb.close()

        return invoices

    def get_invoice_by_id(self, invoice_id: str) -> dict | None:
        for inv in self.get_all_invoices():
            if inv["id"] == invoice_id:
                return inv
        return None

    def get_invoices_by_client(self, client_name: str) -> list[dict]:
        target = client_name.strip().lower()
        return[inv for inv in self.get_all_invoices() if inv["client"].lower() == target]

    def get_next_actions(self) -> list[dict]:
        """
        Returns actionable rows sorted by strict priority matrix.
        Excludes paid/disputed tasks or tasks waiting on future dates.
        """
        invoices = self.get_all_invoices()
        actionable =[]
        today_iso = date.today().isoformat()

        priority_map = {
            "escalate_to_legal": 1,
            "send_final_notice": 2,
            "send_urgent_followup": 3,
            "schedule_call": 4,
            "send_friendly_reminder": 5,
            "resolve_contact_details": 7,
            "human_review_required": 8
        }

        for inv in invoices:
            action = inv["next_action"]
            
            # Exclude blocked/future tasks
            if action == "disputed_under_review":
                continue
            if action.startswith("await_payment_") or action.startswith("follow_up_"):
                # Extract date and skip if it's in the future
                date_str = action.split("_")[-1]
                if len(date_str) == 10 and date_str > today_iso:
                    continue
                # If date is today or past, treat it as priority 6
                inv["_sort_priority"] = 6
                actionable.append(inv)
                continue
                
            if action in priority_map:
                inv["_sort_priority"] = priority_map[action]
                actionable.append(inv)

        return sorted(actionable, key=lambda x: x["_sort_priority"])

    def update_next_action(self, invoice_id: str, action: str, agent_name: str) -> bool:
        """
        Safely writes task queue instructions. Enforces protected fields.
        """
        with locked_excel(self.filepath, mode="w") as wb:
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value) == invoice_id:
                    current_action = str(ws.cell(row=row, column=15).value or "").strip()
                    
                    # Security logic
                    is_protected = (
                        current_action in PROTECTED_ACTIONS or 
                        current_action.startswith("await_payment_")
                    )
                    
                    if is_protected and agent_name != "human":
                        _log_lineage({
                            "agent": "excel_tool",
                            "action": "update_next_action_BLOCKED",
                            "invoice_id": invoice_id,
                            "reason": f"Next Action is '{current_action}' — cannot be overwritten by non-human agent",
                            "attempted_by": agent_name,
                            "write_success": False
                        })
                        return False
                    
                    # Proceed with update
                    ws.cell(row=row, column=15).value = action
                    ws.cell(row=row, column=16).value = agent_name
                    ws.cell(row=row, column=17).value = datetime.now(timezone.utc).isoformat()
                    
                    _log_lineage({
                        "agent": "excel_tool",
                        "action": "update_next_action",
                        "invoice_id": invoice_id,
                        "previous_value": current_action,
                        "new_value": action,
                        "updated_by": agent_name,
                        "write_success": True
                    })
                    return True
        return False

    def update_contact_info(self, invoice_id: str, contact_name: str, contact_phone: str | None) -> bool:
        """
        ATOMIC TWO-FILE SYNC: Updates both Excel and mock_invoices.json.
        Rolls back Excel if JSON write fails.
        """
        with locked_excel(self.filepath, mode="w") as wb:
            ws = wb.active
            target_row = None
            old_name = None
            
            # Step 1: Find and Update Excel internally
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value) == invoice_id:
                    target_row = row
                    old_name = ws.cell(row=row, column=7).value
                    ws.cell(row=row, column=7).value = contact_name
                    # Note: Excel schema doesn't have phone, so we only update name here.
                    break
                    
            if not target_row:
                return False
                
            # Step 2: Attempt JSON Sync
            try:
                with open(JSON_PATH, "r", encoding="utf-8") as f:
                    json_data = json.load(f)
                    
                updated = False
                for item in json_data:
                    if item.get("invoice_id") == invoice_id or item.get("id") == invoice_id:
                        if "contact_info" in item:
                            item["contact_info"]["name"] = contact_name
                            if contact_phone:
                                item["contact_info"]["phone"] = contact_phone
                        else:
                            item["contact_name"] = contact_name
                            if contact_phone:
                                item["contact_phone"] = contact_phone
                        updated = True
                        break
                        
                if not updated:
                    raise ValueError("Invoice ID not found in JSON data")

                with open(JSON_PATH, "w", encoding="utf-8") as f:
                    json.dump(json_data, f, indent=2, ensure_ascii=False)

            except Exception as e:
                # Step 3: ROLLBACK Excel if JSON fails
                ws.cell(row=target_row, column=7).value = old_name
                print(f"[EXCEL SYNC ERROR] Rolling back. {e}")
                return False

            # Lock is released and Excel is saved ONLY if we reach here successfully
            _log_lineage({
                "agent": "excel_tool",
                "action": "update_contact_info",
                "invoice_id": invoice_id,
                "new_contact_name": contact_name,
                "write_success": True
            })
            return True

    def log_contact_made(self, invoice_id: str, contact_type: str, outcome: str, agent_name: str) -> bool:
        """Updates contact history metrics for the dashboard and agent context."""
        with locked_excel(self.filepath, mode="w") as wb:
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value) == invoice_id:
                    today_str = date.today().isoformat()
                    current_count = int(ws.cell(row=row, column=14).value or 0)
                    
                    ws.cell(row=row, column=12).value = today_str          # Last Contact Date
                    ws.cell(row=row, column=13).value = contact_type       # Last Contact Type
                    
                    # If they didn't answer, don't increment the harassment counter
                    if outcome != "no_response":
                        ws.cell(row=row, column=14).value = current_count + 1  
                    
                    ws.cell(row=row, column=16).value = agent_name
                    ws.cell(row=row, column=17).value = datetime.now(timezone.utc).isoformat()
                    return True
        return False

    def mark_paid(self, invoice_id: str) -> bool:
        """Irreversible status closure."""
        with locked_excel(self.filepath, mode="w") as wb:
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value) == invoice_id:
                    ws.cell(row=row, column=6).value = "paid"
                    ws.cell(row=row, column=9).value = 0            # Clear Risk Score
                    ws.cell(row=row, column=10).value = "Low"       # Clear Risk Label
                    ws.cell(row=row, column=15).value = "paid"      # Next Action
                    ws.cell(row=row, column=16).value = "human"
                    ws.cell(row=row, column=17).value = datetime.now(timezone.utc).isoformat()
                    
                    _log_lineage({
                        "agent": "excel_tool",
                        "action": "mark_paid",
                        "invoice_id": invoice_id,
                        "write_success": True
                    })
                    return True
        return False

    def get_overdue_summary(self) -> dict:
        """Dashboard aggregation reading directly from live data."""
        invoices = self.get_all_invoices()
        summary = {
            "total_overdue_amount": sum(inv["amount"] for inv in invoices),
            "count_high_risk": sum(1 for inv in invoices if inv["risk_label"] == "High"),
            "count_medium_risk": sum(1 for inv in invoices if inv["risk_label"] == "Medium"),
            "count_low_risk": sum(1 for inv in invoices if inv["risk_label"] == "Low"),
            "actions_queued": {}
        }
        
        for inv in invoices:
            action = inv["next_action"]
            if action:
                summary["actions_queued"][action] = summary["actions_queued"].get(action, 0) + 1
                
        return summary

    def get_call_script_branches(self, invoice: dict) -> dict:
        """
        Generates the standard Verification branch strings based on invoice context.
        Used by the Voice Agent to strictly enforce right-person verification.
        """
        company = "Apex Collections" # Substitute with actual client firm name in prod
        client = invoice.get("client", "the client")
        contact_name = invoice.get("contact_name", "the account manager")
        amount = f"₹{invoice.get('amount', 0):,.0f}"
        due_date = invoice.get("due_date", "the due date")
        days = invoice.get("days_overdue", 0)
        inv_id = invoice.get("id", "the invoice")

        return {
            "step_1_verification": (
                f"Hello, this is an automated call from {company}'s accounts team. "
                f"May I please speak with {contact_name}?"
            ),
            "branch_confirmed": (
                f"Thank you {contact_name}. I'm calling regarding invoice {inv_id} "
                f"for {amount}, which was due on {due_date} and is now {days} days overdue. "
                f"Could you help us understand when we can expect payment?"
            ),
            "branch_wrong_person": (
                f"I apologize for the interruption. Could you let me know if {contact_name} "
                f"is available, or the best way to reach them?"
            ),
            "branch_wrong_number": (
                f"I sincerely apologize for the inconvenience. It seems we may have "
                f"the wrong number on file. We'll correct this immediately. "
                f"Sorry for disturbing you, have a good day."
            ),
            "branch_graceful_exit": (
                "Thank you for letting me know. We'll try again at a better time. Have a good day."
            ),
            "branch_no_response_retry": "Hello? Can you hear me?"
        }

# Export singleton instance
excel_tool = ExcelTool()