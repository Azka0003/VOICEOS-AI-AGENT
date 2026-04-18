import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("mock_invoices.json", "r") as f:
    invoices = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Invoices"

headers = [
    "Invoice ID", "Client", "Amount (₹)", "Due Date", "Days Overdue",
    "Status", "Contact Name", "Contact Email", "Risk Score", "Risk Label",
    "Dispute", "Next Action"
]

def get_next_action(inv):
    if inv["contact_name"] is None:
        return "Resolve contact details before proceeding"
    if inv["risk_label"] == "High" and inv["days_overdue"] >= 60:
        return "Escalate to legal team"
    if inv["risk_label"] == "Medium" and 30 <= inv["days_overdue"] <= 60:
        return "Schedule follow-up call"
    if inv["risk_label"] == "Low" and inv["days_overdue"] < 30:
        return "Send friendly reminder email"
    if inv["risk_label"] == "High":
        return "Escalate to legal team"
    if inv["risk_label"] == "Medium":
        return "Schedule follow-up call"
    return "Send friendly reminder email"

header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="CCCCCC")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = cell_border

ws.row_dimensions[1].height = 30

fill_low = PatternFill("solid", fgColor="C6EFCE")
fill_medium = PatternFill("solid", fgColor="FFEB9C")
fill_high = PatternFill("solid", fgColor="FFC7CE")
font_low = Font(name="Arial", color="276221", bold=True, size=10)
font_medium = Font(name="Arial", color="9C5700", bold=True, size=10)
font_high = Font(name="Arial", color="9C0006", bold=True, size=10)

alt_fill = PatternFill("solid", fgColor="F2F7FC")
default_fill = PatternFill("solid", fgColor="FFFFFF")

for row_idx, inv in enumerate(invoices, 2):
    row_fill = alt_fill if row_idx % 2 == 0 else default_fill

    data = [
        inv["id"],
        inv["client"],
        inv["amount"],
        inv["due_date"],
        inv["days_overdue"],
        inv["status"].capitalize(),
        inv["contact_name"] if inv["contact_name"] else "— MISSING —",
        inv["contact_email"],
        inv["risk_score"],
        inv["risk_label"],
        "Yes" if inv["dispute_flag"] else "No",
        get_next_action(inv)
    ]

    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.border = cell_border
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 12))
        cell.fill = row_fill

    # Amount formatting
    amt_cell = ws.cell(row=row_idx, column=3)
    amt_cell.number_format = '₹#,##0'

    # Risk Label coloring
    risk_cell = ws.cell(row=row_idx, column=10)
    label = inv["risk_label"]
    if label == "Low":
        risk_cell.fill = fill_low
        risk_cell.font = font_low
    elif label == "Medium":
        risk_cell.fill = fill_medium
        risk_cell.font = font_medium
    elif label == "High":
        risk_cell.fill = fill_high
        risk_cell.font = font_high
    risk_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Contact name red if missing
    if inv["contact_name"] is None:
        ws.cell(row=row_idx, column=7).font = Font(name="Arial", color="9C0006", bold=True, size=10, italic=True)

    # Days overdue color
    days_cell = ws.cell(row=row_idx, column=5)
    if inv["days_overdue"] >= 60:
        days_cell.font = Font(name="Arial", color="9C0006", bold=True, size=10)
    elif inv["days_overdue"] >= 30:
        days_cell.font = Font(name="Arial", color="9C5700", size=10)

    # Dispute flag highlight
    dispute_cell = ws.cell(row=row_idx, column=11)
    if inv["dispute_flag"]:
        dispute_cell.fill = fill_high
        dispute_cell.font = Font(name="Arial", color="9C0006", bold=True, size=10)
    dispute_cell.alignment = Alignment(horizontal="center", vertical="center")

col_widths = [12, 20, 14, 13, 14, 10, 18, 32, 11, 12, 10, 40]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(invoices)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "VoiceOS – Invoice Portfolio Summary"
ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
ws2["A1"].alignment = Alignment(horizontal="left")

ws2["A3"] = "Metric"
ws2["B3"] = "Value"
for cell in [ws2["A3"], ws2["B3"]]:
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")

metrics = [
    ("Total Invoices", f"=COUNTA(Invoices!A2:A{len(invoices)+1})"),
    ("Total Outstanding (₹)", f"=SUM(Invoices!C2:C{len(invoices)+1})"),
    ("High Risk Invoices", f'=COUNTIF(Invoices!J2:J{len(invoices)+1},"High")'),
    ("Medium Risk Invoices", f'=COUNTIF(Invoices!J2:J{len(invoices)+1},"Medium")'),
    ("Low Risk Invoices", f'=COUNTIF(Invoices!J2:J{len(invoices)+1},"Low")'),
    ("Disputed Invoices", f'=COUNTIF(Invoices!K2:K{len(invoices)+1},"Yes")'),
    ("Missing Contact", f'=COUNTIF(Invoices!G2:G{len(invoices)+1},"— MISSING —")'),
    ("Avg Days Overdue", f"=AVERAGE(Invoices!E2:E{len(invoices)+1})"),
    ("Max Days Overdue", f"=MAX(Invoices!E2:E{len(invoices)+1})"),
]

for i, (label, formula) in enumerate(metrics, 4):
    ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10)
    val_cell = ws2.cell(row=i, column=2, value=formula)
    val_cell.font = Font(name="Arial", size=10)
    if "Outstanding" in label:
        val_cell.number_format = '₹#,##0'
    elif "Avg" in label:
        val_cell.number_format = "0.0"

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 20

wb.save("invoices.xlsx")
print("invoices.xlsx generated successfully with 50 invoice rows and Summary sheet.")
